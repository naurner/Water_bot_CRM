import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import os
from datetime import datetime
from config import USERS_FILE, ORDERS_FILE


class Database:
    """Класс для работы с Excel файлами"""

    @staticmethod
    def _format_headers(ws):
        """Форматирование заголовков листа"""
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    @staticmethod
    def _create_sheet_with_headers(wb, sheet_name, headers):
        """Создание листа с заголовками"""
        ws = wb.create_sheet(sheet_name) if sheet_name not in wb.sheetnames else wb[sheet_name]
        if ws.max_row == 1 or not ws[1][0].value:
            ws.append(headers)
            Database._format_headers(ws)
        return ws

    @staticmethod
    def init_users_file():
        """Инициализация файла пользователей"""
        if not os.path.exists(USERS_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = "Пользователи"
            headers = ['User ID', 'Имя', 'Телефон', 'Адрес', 'Дата регистрации']
            ws.append(headers)
            Database._format_headers(ws)
            wb.save(USERS_FILE)

    @staticmethod
    def init_orders_file():
        """Инициализация файла заказов"""
        if not os.path.exists(ORDERS_FILE):
            wb = Workbook()
            ws = wb.active
            ws.title = datetime.now().strftime('%Y-%m-%d')
            headers = ['Номер заказа', 'User ID', 'Имя', 'Телефон', 'Адрес',
                      'Дата заказа', 'Время доставки', 'Количество бутылок', 'Статус',
                      'Morning Reminder ID', 'Pre-delivery Reminder ID']
            ws.append(headers)
            Database._format_headers(ws)
            wb.save(ORDERS_FILE)

    @staticmethod
    def get_user(user_id):
        """Получить данные пользователя по ID"""
        if not os.path.exists(USERS_FILE):
            return None

        wb = openpyxl.load_workbook(USERS_FILE)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == user_id:
                return {
                    'user_id': row[0],
                    'name': row[1],
                    'phone': row[2],
                    'address': row[3],
                    'registration_date': row[4]
                }
        return None

    @staticmethod
    def save_user(user_id, name, phone, address):
        """Сохранить или обновить данные пользователя"""
        Database.init_users_file()
        wb = openpyxl.load_workbook(USERS_FILE)
        ws = wb.active

        # Ищем существующего пользователя
        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value == user_id:
                ws.cell(idx, 2, name)
                ws.cell(idx, 3, phone)
                ws.cell(idx, 4, address)
                wb.save(USERS_FILE)
                return

        # Добавляем нового пользователя
        ws.append([user_id, name, phone, address, datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        wb.save(USERS_FILE)

    @staticmethod
    def _get_order_headers():
        """Получить заголовки для листа заказов"""
        return ['Номер заказа', 'User ID', 'Имя', 'Телефон', 'Адрес',
                'Дата заказа', 'Время доставки', 'Количество бутылок', 'Статус',
                'Morning Reminder ID', 'Pre-delivery Reminder ID']

    @staticmethod
    def _parse_order_row(row, delivery_date=None):
        """Парсинг строки заказа в словарь"""
        if not row[0]:
            return None

        return {
            'order_id': row[0],
            'user_id': row[1],
            'name': row[2],
            'phone': row[3],
            'address': row[4],
            'order_date': row[5],
            'delivery_time': row[6],
            'bottles': row[7] if len(row) > 7 else 1,
            'status': row[8] if len(row) > 8 else 'Новый',
            'morning_reminder_id': row[9] if len(row) > 9 else None,
            'pre_delivery_reminder_id': row[10] if len(row) > 10 else None,
            'delivery_date': delivery_date
        }

    @staticmethod
    def get_orders_for_date(date_str):
        """Получить все заказы на определенную дату"""
        if not os.path.exists(ORDERS_FILE):
            return []

        wb = openpyxl.load_workbook(ORDERS_FILE)
        if date_str not in wb.sheetnames:
            return []

        ws = wb[date_str]
        orders = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            order = Database._parse_order_row(row, date_str)
            if order:
                orders.append(order)

        return orders

    @staticmethod
    def save_order(user_id, name, phone, address, delivery_date, delivery_time, bottles=1):
        """Сохранить заказ"""
        Database.init_orders_file()
        wb = openpyxl.load_workbook(ORDERS_FILE)
        date_str = delivery_date.strftime('%Y-%m-%d')

        # Создаем или получаем лист для даты
        if date_str not in wb.sheetnames:
            ws = Database._create_sheet_with_headers(wb, date_str, Database._get_order_headers())
        else:
            ws = wb[date_str]

        # Генерируем уникальный ID заказа
        order_id = f"ORD-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        order_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        ws.append([order_id, user_id, name, phone, address, order_date, delivery_time, bottles, "Новый"])
        wb.save(ORDERS_FILE)

        return order_id

    @staticmethod
    def get_user_orders(user_id):
        """Получить все заказы пользователя"""
        if not os.path.exists(ORDERS_FILE):
            return []

        wb = openpyxl.load_workbook(ORDERS_FILE)
        user_orders = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1] == user_id:
                    order = Database._parse_order_row(row, sheet_name)
                    if order:
                        user_orders.append(order)

        # Сортируем по дате доставки
        user_orders.sort(key=lambda x: (x['delivery_date'], x['delivery_time']))
        return user_orders

    @staticmethod
    def get_active_user_orders(user_id):
        """Получить только активные (будущие) заказы пользователя"""
        all_orders = Database.get_user_orders(user_id)
        active_orders = []
        now = datetime.now()

        for order in all_orders:
            if order.get('status') == 'Отменен':
                continue

            try:
                order_datetime = datetime.strptime(
                    f"{order['delivery_date']} {order['delivery_time']}",
                    '%Y-%m-%d %H:%M'
                )
                if order_datetime > now:
                    active_orders.append(order)
            except:
                continue

        return active_orders

    @staticmethod
    def _find_and_delete_order(order_id):
        """Найти и удалить заказ из базы"""
        if not os.path.exists(ORDERS_FILE):
            return False

        wb = openpyxl.load_workbook(ORDERS_FILE)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if row[0].value == order_id:
                    ws.delete_rows(idx)
                    wb.save(ORDERS_FILE)
                    return True

        return False

    @staticmethod
    def cancel_order(order_id):
        """Отменить заказ (удалить из базы, чтобы освободить время)"""
        return Database._find_and_delete_order(order_id)

    @staticmethod
    def delete_order(order_id):
        """Удалить заказ (алиас для cancel_order)"""
        return Database._find_and_delete_order(order_id)

    @staticmethod
    def get_order_by_id(order_id):
        """Получить заказ по ID"""
        if not os.path.exists(ORDERS_FILE):
            return None

        wb = openpyxl.load_workbook(ORDERS_FILE)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == order_id:
                    return Database._parse_order_row(row, sheet_name)

        return None

    @staticmethod
    def is_time_slot_available(date_str, time_str):
        """Проверить, доступен ли временной слот (интервал 30 минут)"""
        orders = Database.get_orders_for_date(date_str)
        if not orders:
            return True

        requested_time = datetime.strptime(time_str, '%H:%M')

        for order in orders:
            if order['delivery_time']:
                order_time = datetime.strptime(order['delivery_time'], '%H:%M')
                time_diff = abs((requested_time - order_time).total_seconds() / 60)
                if time_diff < 30:
                    return False

        return True

    @staticmethod
    def reschedule_order(order_id, new_date_str, new_time_str):
        """Перенести заказ на новую дату и время"""
        order = Database.get_order_by_id(order_id)
        if not order or not Database._find_and_delete_order(order_id):
            return False

        wb = openpyxl.load_workbook(ORDERS_FILE)

        # Создаем или получаем лист для новой даты
        if new_date_str not in wb.sheetnames:
            ws = Database._create_sheet_with_headers(wb, new_date_str, Database._get_order_headers())
        else:
            ws = wb[new_date_str]

        # Добавляем заказ с новым временем
        ws.append([
            order['order_id'],
            order['user_id'],
            order['name'],
            order['phone'],
            order['address'],
            order['order_date'],
            new_time_str,
            order['bottles'],
            'Перенесен'
        ])

        wb.save(ORDERS_FILE)
        return True

    @staticmethod
    def update_order_schedule(order_id, new_date_str, new_time_str):
        """Обновить дату и время заказа (алиас для reschedule_order)"""
        return Database.reschedule_order(order_id, new_date_str, new_time_str)

    @staticmethod
    def _update_user_field(user_id, field_index, value):
        """Обновить поле пользователя по индексу"""
        if not os.path.exists(USERS_FILE):
            return False

        wb = openpyxl.load_workbook(USERS_FILE)
        ws = wb.active

        for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if row[0].value == user_id:
                ws.cell(idx, field_index, value)
                wb.save(USERS_FILE)
                return True

        return False

    @staticmethod
    def update_user_name(user_id, new_name):
        """Обновить имя пользователя"""
        return Database._update_user_field(user_id, 2, new_name)

    @staticmethod
    def update_user_phone(user_id, new_phone):
        """Обновить номер телефона пользователя"""
        from utils import format_kyrgyzstan_phone
        return Database._update_user_field(user_id, 3, format_kyrgyzstan_phone(new_phone))

    @staticmethod
    def update_user_address(user_id, new_address):
        """Обновить адрес пользователя"""
        return Database._update_user_field(user_id, 4, new_address)

    @staticmethod
    def update_order_reminder_ids(order_id, morning_msg_id, pre_delivery_msg_id):
        """
        Обновить ID запланированных сообщений для заказа

        :param order_id: ID заказа
        :param morning_msg_id: ID утреннего напоминания
        :param pre_delivery_msg_id: ID напоминания за 30 минут
        :return: True если успешно, False если нет
        """
        if not os.path.exists(ORDERS_FILE):
            return False

        wb = openpyxl.load_workbook(ORDERS_FILE)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if row[0].value == order_id:
                    # Обновляем ID сообщений (колонки 10 и 11)
                    ws.cell(idx, 10, morning_msg_id)
                    ws.cell(idx, 11, pre_delivery_msg_id)
                    wb.save(ORDERS_FILE)
                    return True

        return False

    @staticmethod
    def get_order_reminder_ids(order_id):
        """
        Получить ID запланированных сообщений для заказа

        :param order_id: ID заказа
        :return: Tuple (morning_msg_id, pre_delivery_msg_id) или (None, None)
        """
        order = Database.get_order_by_id(order_id)
        if order:
            return (
                order.get('morning_reminder_id'),
                order.get('pre_delivery_reminder_id')
            )
        return (None, None)
